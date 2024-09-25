# Import package
import openpyxl

# Load Workbook
wk = openpyxl.load_workbook("C:\\Users\\a315707\\OneDrive - Deutsche Telekom AG\\Schulungen\\Programmierung\\Python\\PythonLearning\\18_ExcelData\\TestData.xlsx")

#print(wk.sheetnames) # List sheets
#print(wk.active.title) # Active Sheet title

# Create object of sheet on which you want to work

sh = wk['Tabelle1'] # Move to Workbook level
#print(sh.title) # Active Sheet title
print(sh['A3'].value)

c1=sh.cell(1,1) # Move to cell level - first row, first column
print(c1.value)

c2=sh.cell(column=3, row=2) # without these arguments, it's first row then column
print(c2.value)

# Read and print all the data
#rows=sh.max_row
#columns=sh.max_column
#print("Total rows are " + str(rows))
#print("Total columns are " + str(columns))
#for i in range(1,rows+1):
#    for j in range(1,columns+1):
#        print(str(sh.cell(i,j).value))

# Alternative approach
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)



