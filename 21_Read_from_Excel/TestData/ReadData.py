import  openpyxl

wk = openpyxl.load_workbook("C://Users//a315707\OneDrive - Deutsche Telekom AG//Schulungen//Programmierung//Python//PythonLearning//21_Read_from_Excel//TestData//TestData100.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(row,cell)
    return cell.value

print(fetch_number_of_rows("Tabelle1"))
print(fetch_cell_data("Tabelle1",1,1))